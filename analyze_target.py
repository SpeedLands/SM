import openpyxl

file_path = "Datos transformados .xlsx"
wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    # Print first row of data
    for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        print(f"Sample data: {row}")
