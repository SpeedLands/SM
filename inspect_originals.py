import openpyxl
import os

files = [
    "DIRECTORIO ALUMNOS 3°A  T.M..xlsm",
    "DIRECTORIO ALUMNOS 3°B  T.M..xlsm",
    "DIRECTORIO ALUMNOS 3°C T.M..xlsm",
    "DIRECTORIO ALUMNOS 3°D T.M..xlsm"
]

for file in files:
    print(f"\n--- Analyzing file: {file} ---")
    if not os.path.exists(file):
        print("File not found.")
        continue
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        # Check first sheet
        sheet = wb.active # Usually the first or "Hoja1"
        print(f"Active Sheet: {sheet.title}")
        
        # Get headers (first few rows might be headers or titles)
        print("First 10 rows:")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            print(f"Row {i}: {row}")
            
    except Exception as e:
        print(f"Error reading file: {e}")
